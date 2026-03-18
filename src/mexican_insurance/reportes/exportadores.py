"""
Exportadores de reportes a Excel y CSV.

Convierte DataFrames de reportes a archivos Excel formateados
o CSV para integración con otros sistemas.
"""

from datetime import datetime
from pathlib import Path

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportadorExcel:
    """
    Exporta reportes a archivos Excel con formato profesional.

    Ejemplo:
        >>> exportador = ExportadorExcel()
        >>> exportador.exportar_reporte_completo(
        ...     ruta_salida="reporte_Q1_2024.xlsx",
        ...     df_suscripcion=df_susc,
        ...     df_siniestros=df_sin,
        ...     df_inversiones=df_inv,
        ...     df_rcs=df_rcs,
        ...     metadata=metadata
        ... )
    """

    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl no está instalado. Instálalo con: pip install openpyxl"
            )

    def exportar_reporte_completo(
        self,
        ruta_salida: str,
        df_suscripcion: pd.DataFrame | None = None,
        df_siniestros: pd.DataFrame | None = None,
        df_inversiones: pd.DataFrame | None = None,
        df_rcs: pd.DataFrame | None = None,
        metadata: dict | None = None,
    ) -> Path:
        """
        Exporta reporte completo a Excel con múltiples hojas.

        Args:
            ruta_salida: Ruta del archivo Excel de salida
            df_suscripcion: DataFrame de suscripción (opcional)
            df_siniestros: DataFrame de siniestros (opcional)
            df_inversiones: DataFrame de inversiones (opcional)
            df_rcs: DataFrame de RCS (opcional)
            metadata: Metadatos del reporte (opcional)

        Returns:
            Path del archivo generado
        """
        ruta = Path(ruta_salida)

        # Crear Excel writer
        with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
            # Hoja de portada
            if metadata:
                df_portada = self._crear_hoja_portada(metadata)
                df_portada.to_excel(writer, sheet_name="Portada", index=False)

            # Hojas de datos
            if df_suscripcion is not None and not df_suscripcion.empty:
                df_suscripcion.to_excel(
                    writer, sheet_name="Suscripción", index=False
                )

            if df_siniestros is not None and not df_siniestros.empty:
                df_siniestros.to_excel(writer, sheet_name="Siniestros", index=False)

            if df_inversiones is not None and not df_inversiones.empty:
                df_inversiones.to_excel(writer, sheet_name="Inversiones", index=False)

            if df_rcs is not None and not df_rcs.empty:
                df_rcs.to_excel(writer, sheet_name="RCS", index=False)

        # Aplicar formato
        self._aplicar_formato(ruta)

        return ruta

    def _crear_hoja_portada(self, metadata: dict) -> pd.DataFrame:
        """Crea DataFrame para hoja de portada"""
        datos = [
            ["REPORTE TRIMESTRAL CNSF", ""],
            ["", ""],
            ["Información de la Aseguradora", ""],
            ["Clave", metadata.get("clave_aseguradora", "")],
            ["Nombre", metadata.get("nombre_aseguradora", "")],
            ["", ""],
            ["Información del Período", ""],
            ["Trimestre", metadata.get("trimestre", "")],
            ["Año", metadata.get("anio", "")],
            ["Fecha de Presentación", str(metadata.get("fecha_presentacion", ""))],
            ["", ""],
            ["Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ]

        return pd.DataFrame(datos, columns=["Campo", "Valor"])

    def _aplicar_formato(self, ruta: Path) -> None:
        """Aplica formato profesional al archivo Excel"""
        if not OPENPYXL_AVAILABLE:
            return

        wb = load_workbook(ruta)

        # Estilos
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Formato de encabezados (primera fila)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Formato de números
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        # Si es número grande (>1000), formato con separador de miles
                        if abs(cell.value) >= 1000:
                            cell.number_format = "#,##0.00"
                        elif 0 < abs(cell.value) < 1:
                            # Si es porcentaje pequeño
                            cell.number_format = "0.00"
                        else:
                            cell.number_format = "0.00"

            # Congelar primera fila
            ws.freeze_panes = "A2"

        wb.save(ruta)


class ExportadorCSV:
    """
    Exporta reportes a archivos CSV.

    Ejemplo:
        >>> exportador = ExportadorCSV()
        >>> exportador.exportar_dataframe(
        ...     df=df_suscripcion,
        ...     ruta_salida="suscripcion_Q1_2024.csv"
        ... )
    """

    def exportar_dataframe(
        self, df: pd.DataFrame, ruta_salida: str, separador: str = ","
    ) -> Path:
        """
        Exporta DataFrame a CSV.

        Args:
            df: DataFrame a exportar
            ruta_salida: Ruta del archivo CSV de salida
            separador: Separador de columnas (default: ',')

        Returns:
            Path del archivo generado
        """
        ruta = Path(ruta_salida)
        df.to_csv(ruta, sep=separador, index=False, encoding="utf-8-sig")
        return ruta

    def exportar_multiples(
        self,
        dataframes: dict[str, pd.DataFrame],
        directorio_salida: str,
        prefijo: str = "",
        separador: str = ",",
    ) -> list[Path]:
        """
        Exporta múltiples DataFrames a archivos CSV separados.

        Args:
            dataframes: Diccionario {nombre: DataFrame}
            directorio_salida: Directorio donde guardar los archivos
            prefijo: Prefijo para nombres de archivo (opcional)
            separador: Separador de columnas (default: ',')

        Returns:
            Lista de rutas de archivos generados
        """
        directorio = Path(directorio_salida)
        directorio.mkdir(parents=True, exist_ok=True)

        rutas = []
        for nombre, df in dataframes.items():
            if df is not None and not df.empty:
                nombre_archivo = f"{prefijo}{nombre}.csv" if prefijo else f"{nombre}.csv"
                ruta = directorio / nombre_archivo
                self.exportar_dataframe(df, str(ruta), separador)
                rutas.append(ruta)

        return rutas
